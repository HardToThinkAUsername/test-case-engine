"""Excel导出工具。

将测试用例导出为格式化的Excel文档，包含：
- Sheet 1: 测试用例明细（全量用例，冻结首行，自动筛选，优先级着色）
- Sheet 2: 统计汇总（优先级分布、模块分布、类型分布、来源分布）
- Sheet 3: 按模块（用例按模块分组展示）
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 列定义 ──
COLUMNS = [
    {"header": "用例编号", "key": "id", "width": 18},
    {"header": "用例标题", "key": "title", "width": 44},
    {"header": "优先级", "key": "priority", "width": 8},
    {"header": "用例类型", "key": "type", "width": 10},
    {"header": "测试类型", "key": "test_type", "width": 10},
    {"header": "所属模块", "key": "module", "width": 18},
    {"header": "来源", "key": "source", "width": 22},
    {"header": "前置条件", "key": "preconditions", "width": 35},
    {"header": "操作步骤", "key": "steps", "width": 55},
    {"header": "预期结果", "key": "expected_result", "width": 45},
    {"header": "关联需求", "key": "related_requirement", "width": 20},
]

# ── 样式 ──
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(bold=True, size=12, color="4472C4")
MODULE_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MODULE_HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUB_HEADER_FONT = Font(bold=True, size=10, color="1F4E79")
P1_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
P2_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
P3_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
P1_FONT = Font(color="C62828", bold=True)
P2_FONT = Font(color="E65100", bold=True)
P3_FONT = Font(color="2E7D32", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top")

# 优先级 → 样式映射
PRIORITY_STYLE = {
    "P1": (P1_FILL, P1_FONT),
    "P2": (P2_FILL, P2_FONT),
    "P3": (P3_FILL, P3_FONT),
}

# 类型中文映射
TYPE_LABELS = {
    "positive": "正向",
    "branch": "分支",
    "exception": "异常",
    "boundary": "边界",
}
TEST_TYPE_LABELS = {
    "functional": "功能测试",
    "performance": "性能测试",
}


def _format_list(value: Any) -> str:
    """将列表类型转换为带序号的换行字符串。"""
    if isinstance(value, list):
        return "\n".join(f"{i+1}. {item}" for i, item in enumerate(value))
    return str(value) if value is not None else ""


def _write_header_row(ws, headers: list[str], row: int = 1) -> None:
    """写入表头行。"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _set_column_widths(ws, widths: list[int]) -> None:
    """设置列宽。"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


# ── Sheet 1: 测试用例明细 ──
def _write_detail_sheet(wb: Workbook, test_cases: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "测试用例明细"

    # 表头
    headers = [c["header"] for c in COLUMNS]
    _write_header_row(ws, headers)
    ws.freeze_panes = "A2"

    # 数据
    for row_idx, case in enumerate(test_cases, 2):
        for col_idx, col_def in enumerate(COLUMNS, 1):
            value = _format_list(case.get(col_def["key"], ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGNMENT

            # 优先级着色
            priority = case.get("priority", "")
            if priority in PRIORITY_STYLE:
                fill, font = PRIORITY_STYLE[priority]
                if col_idx == 3:  # 优先级列
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = CENTER_ALIGNMENT

        ws.row_dimensions[row_idx].height = 72

    # 列宽 + 自动筛选
    _set_column_widths(ws, [c["width"] for c in COLUMNS])
    ws.auto_filter.ref = ws.dimensions


# ── Sheet 2: 统计汇总 ──
def _write_summary_sheet(wb: Workbook, test_cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("统计汇总")

    total = len(test_cases)

    # 标题
    ws.cell(row=1, column=1, value="测试用例统计汇总").font = TITLE_FONT
    ws.merge_cells("A1:C1")

    row = 3

    # ── 优先级分布 ──
    ws.cell(row=row, column=1, value="优先级分布").font = SECTION_FONT
    row += 1
    _write_header_row(ws, ["优先级", "数量", "占比"], row)
    row += 1
    priority_counter = Counter(c.get("priority", "") for c in test_cases)
    for p in ["P1", "P2", "P3"]:
        count = priority_counter.get(p, 0)
        ws.cell(row=row, column=1, value=p).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%").border = THIN_BORDER
        fill, font = PRIORITY_STYLE.get(p, (None, None))
        if fill:
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=1).font = font
        row += 1
    ws.cell(row=row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=total).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3, value="100%").border = THIN_BORDER
    row += 2

    # ── 模块分布 ──
    ws.cell(row=row, column=1, value="模块分布").font = SECTION_FONT
    row += 1
    _write_header_row(ws, ["模块", "数量", "占比"], row)
    row += 1
    module_counter = Counter(c.get("module", "") for c in test_cases)
    for mod, count in module_counter.most_common():
        ws.cell(row=row, column=1, value=mod).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%").border = THIN_BORDER
        row += 1
    row += 1

    # ── 用例类型分布 ──
    ws.cell(row=row, column=1, value="用例类型分布").font = SECTION_FONT
    row += 1
    _write_header_row(ws, ["类型", "数量", "占比"], row)
    row += 1
    type_counter = Counter(c.get("type", "") for c in test_cases)
    for t, count in type_counter.most_common():
        label = TYPE_LABELS.get(t, t)
        ws.cell(row=row, column=1, value=f"{label} ({t})").border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%").border = THIN_BORDER
        row += 1
    row += 1

    # ── 来源分布 ──
    ws.cell(row=row, column=1, value="来源层级分布").font = SECTION_FONT
    row += 1
    _write_header_row(ws, ["来源层级", "数量", "占比"], row)
    row += 1
    source_counter: Counter[str] = Counter()
    for c in test_cases:
        src = c.get("source", "")
        if src.startswith("L1"):
            source_counter["L1 主路径"] += 1
        elif src.startswith("L2"):
            source_counter["L2 分支路径"] += 1
        elif src.startswith("L3"):
            source_counter["L3 异常模板"] += 1
        elif src.startswith("L4"):
            source_counter["L4 想象力场景"] += 1
        else:
            source_counter[src] += 1
    for src, count in source_counter.most_common():
        ws.cell(row=row, column=1, value=src).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%").border = THIN_BORDER
        row += 1

    _set_column_widths(ws, [24, 12, 12])


# ── Sheet 3: 按模块 ──
def _write_module_sheet(wb: Workbook, test_cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("按模块")

    # 按模块分组
    modules: dict[str, list[dict[str, Any]]] = {}
    for tc in test_cases:
        mod = tc.get("module", "未分类")
        modules.setdefault(mod, []).append(tc)

    # 精简列
    slim_headers = ["用例编号", "用例标题", "优先级", "类型", "来源", "预期结果"]
    slim_keys = ["id", "title", "priority", "type", "source", "expected_result"]
    slim_widths = [18, 44, 8, 10, 22, 45]

    row = 1
    for mod_name, cases in modules.items():
        # 模块标题行
        cell = ws.cell(row=row, column=1, value=mod_name)
        cell.font = MODULE_HEADER_FONT
        cell.fill = MODULE_HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, len(slim_headers) + 1):
            ws.cell(row=row, column=c).fill = MODULE_HEADER_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(slim_headers))
        row += 1

        # 子表头
        for col_idx, header in enumerate(slim_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = SUB_HEADER_FILL
            cell.font = SUB_HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        row += 1

        # 用例行
        for tc in cases:
            for col_idx, key in enumerate(slim_keys, 1):
                value = _format_list(tc.get(key, ""))
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

                # 优先级着色
                if key == "priority":
                    priority = tc.get("priority", "")
                    if priority in PRIORITY_STYLE:
                        fill, font = PRIORITY_STYLE[priority]
                        cell.fill = fill
                        cell.font = font
                        cell.alignment = CENTER_ALIGNMENT

            ws.row_dimensions[row].height = 48
            row += 1

        row += 1  # 模块间空行

    _set_column_widths(ws, slim_widths)
    ws.freeze_panes = "A2"


# ── 主入口 ──
def export_to_excel(
    test_cases: list[dict[str, Any]],
    output_path: Path | str,
) -> Path:
    """将测试用例导出为多 Sheet 格式化 Excel。

    Args:
        test_cases: 测试用例列表
        output_path: 输出文件路径

    Returns:
        输出文件路径
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _write_detail_sheet(wb, test_cases)
    _write_summary_sheet(wb, test_cases)
    _write_module_sheet(wb, test_cases)

    wb.save(output_path)
    return output_path